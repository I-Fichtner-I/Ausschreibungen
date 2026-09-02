"""XLSX-Extraktion mit openpyxl.

Jedes Arbeitsblatt wird zu einer Tabelle **und** zu einer Textseite: die Tabelle
fuer die Artikelerkennung (Stufe 3), der Text fuer Stichwortsuche und
Anforderungserkennung.
"""

from __future__ import annotations

from pathlib import Path
from typing import ClassVar

from openpyxl import load_workbook

from ..models.document import ExtractedDocument, ExtractedPage, ExtractedTable
from .base import DocumentExtractor, register_extractor

_XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@register_extractor
class XlsxExtractor(DocumentExtractor):
    name: ClassVar[str] = "xlsx"
    media_types: ClassVar[tuple[str, ...]] = (_XLSX_MEDIA, "application/vnd.ms-excel")
    suffixes: ClassVar[tuple[str, ...]] = (".xlsx", ".xlsm")

    def _extract(self, path: Path, document: ExtractedDocument) -> None:
        # data_only: berechnete Werte statt Formeln - fuer Mengen und Preise
        # zaehlt das Ergebnis, nicht die Formel.
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            document.metadata = {"sheets": ", ".join(workbook.sheetnames)}
            for number, sheet in enumerate(workbook.worksheets, start=1):
                rows: list[list[str]] = []
                for raw_row in sheet.iter_rows(values_only=True):
                    row = ["" if cell is None else str(cell).strip() for cell in raw_row]
                    if any(row):
                        rows.append(row)
                if not rows:
                    document.pages.append(ExtractedPage(number=number, text=""))
                    continue

                header = rows[0] if len(rows) > 1 and all(cell for cell in rows[0]) else None
                document.tables.append(
                    ExtractedTable(
                        page=number,
                        index=0,
                        section=sheet.title,
                        header=header,
                        rows=rows[1:] if header else rows,
                    )
                )
                text_rows = [" | ".join(cell for cell in row if cell) for row in rows]
                document.pages.append(
                    ExtractedPage(number=number, text="\n".join(filter(None, text_rows)))
                )
        finally:
            workbook.close()
