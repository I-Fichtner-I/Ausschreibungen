"""Lieferantenpreisliste als Preisquelle.

Der praktisch wichtigste Fall - und der einzige, der ohne fremde Infrastruktur
auskommt: Haendler bekommen von ihren Lieferanten Preislisten als CSV, XLSX
oder JSON. Diese Quelle liest sie ein und macht daraus durchsuchbare Angebote.

Das ist kein Notbehelf gegenueber einer Shop-Recherche, sondern die
belastbarere Grundlage: der Einkaufspreis aus der eigenen Lieferantenliste ist
der Preis, zu dem tatsaechlich beschafft wird - ein Schaufensterpreis ist es
nicht.

Die Spaltenzuordnung ist konfigurierbar, weil jede Liste anders aussieht.
Fehlt eine Angabe (Netto/Brutto, Versand, Verfuegbarkeit), bleibt sie leer und
schlaegt spaeter als "nicht kalkulationsfaehig" durch - sie wird nicht geraten.
"""

from __future__ import annotations

import csv
import json
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from ...config import CatalogPriceSourceConfig
from ...core.errors import ConfigError
from ...models.common import Provenance
from ...models.price import Availability, PriceBasis, PriceQuote, PriceTier
from ...sources.parsing import parse_amount, parse_datetime
from ..matching import normalize_identifier, title_overlap
from .base import PriceSource, PriceSourceStatus, ProductQuery
from .registry import register_price_source

#: Schreibweisen der Bezugsgroesse in Preislisten.
_BASIS_VALUES = {
    "net": PriceBasis.NET,
    "netto": PriceBasis.NET,
    "ohne mwst": PriceBasis.NET,
    "exkl": PriceBasis.NET,
    "gross": PriceBasis.GROSS,
    "brutto": PriceBasis.GROSS,
    "inkl mwst": PriceBasis.GROSS,
}
_AVAILABILITY_VALUES = {
    "lager": Availability.IN_STOCK,
    "auf lager": Availability.IN_STOCK,
    "vorraetig": Availability.IN_STOCK,
    "verfuegbar": Availability.IN_STOCK,
    "in stock": Availability.IN_STOCK,
    "bestellbar": Availability.ON_ORDER,
    "auf bestellung": Availability.ON_ORDER,
    "on order": Availability.ON_ORDER,
    "nicht verfuegbar": Availability.UNAVAILABLE,
    "ausgelaufen": Availability.UNAVAILABLE,
    "unavailable": Availability.UNAVAILABLE,
}


def _clean(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize_key(value: str | None) -> str:
    if not value:
        return ""
    return " ".join(value.casefold().replace(".", " ").split())


def _parse_basis(value: str | None) -> PriceBasis:
    """Unbekannte oder fehlende Angabe bleibt UNKNOWN - nie geraten."""
    key = _normalize_key(value)
    if not key:
        return PriceBasis.UNKNOWN
    for token, basis in _BASIS_VALUES.items():
        if token in key:
            return basis
    return PriceBasis.UNKNOWN


def _parse_availability(value: str | None) -> Availability:
    key = _normalize_key(value)
    if not key:
        return Availability.UNKNOWN
    for token, availability in _AVAILABILITY_VALUES.items():
        if token in key:
            return availability
    return Availability.UNKNOWN


def _parse_rate(value: Any) -> float | None:
    """Steuersatz als Anteil: "19", "19 %" und "0,19" ergeben 0.19."""
    amount = parse_amount(value)
    if amount is None:
        return None
    if amount < 0:
        return None
    return amount / 100.0 if amount > 1.0 else amount


def _parse_tiers(value: Any) -> list[PriceTier]:
    """Staffeln als "50:89,90; 100:84,50"."""
    text = _clean(value)
    if not text:
        return []
    tiers: list[PriceTier] = []
    for part in text.replace("|", ";").split(";"):
        if ":" not in part:
            continue
        quantity_text, amount_text = part.split(":", 1)
        quantity = parse_amount(quantity_text)
        amount = parse_amount(amount_text)
        if quantity is None or amount is None:
            continue
        tiers.append(PriceTier(min_quantity=quantity, amount=amount))
    return sorted(tiers, key=lambda tier: tier.min_quantity)


@register_price_source
class CatalogPriceSource(PriceSource):
    """Preisliste aus CSV, XLSX oder JSON."""

    type_name = "catalog"

    config: CatalogPriceSourceConfig

    def __init__(self, *args: Any, **kwargs: Any) -> None:
        super().__init__(*args, **kwargs)
        self._rows: list[dict[str, Any]] | None = None
        self._loaded_at: datetime | None = None

    # -- Laden ---------------------------------------------------------------

    @property
    def path(self) -> Path:
        return Path(self.config.path).expanduser()

    def _load_rows(self) -> list[dict[str, Any]]:
        """Preisliste einlesen; das Ergebnis wird je Lauf einmal gehalten."""
        if self._rows is not None:
            return self._rows
        path = self.path
        if not path.exists():
            raise ConfigError(f"Preisliste nicht gefunden: {path}")

        suffix = path.suffix.casefold()
        if suffix == ".json":
            rows = self._load_json(path)
        elif suffix in (".xlsx", ".xlsm"):
            rows = self._load_xlsx(path)
        else:
            rows = self._load_csv(path)

        self._rows = rows
        self._loaded_at = datetime.now(UTC)
        self.log.info("catalog_loaded", path=str(path), rows=len(rows))
        return rows

    def _load_json(self, path: Path) -> list[dict[str, Any]]:
        payload = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(payload, dict):
            for key in ("products", "items", "rows", "articles"):
                if isinstance(payload.get(key), list):
                    payload = payload[key]
                    break
            else:
                raise ConfigError(
                    f"{path}: erwartet wurde eine Liste oder ein Objekt mit "
                    f"'products'/'items'/'rows'/'articles'"
                )
        if not isinstance(payload, list):
            raise ConfigError(f"{path}: erwartet wurde eine Liste von Artikeln")
        return [row for row in payload if isinstance(row, dict)]

    def _load_csv(self, path: Path) -> list[dict[str, Any]]:
        text = path.read_text(encoding=self.config.encoding, errors="replace")
        sample = text[:4096]
        try:
            dialect: Any = csv.Sniffer().sniff(sample, delimiters=";,\t|")
        except csv.Error:
            # Einspaltige oder ungewoehnliche Datei: Semikolon ist im
            # deutschsprachigen Raum die haeufigere Wahl.
            dialect = csv.excel
            dialect.delimiter = ";"
        return list(csv.DictReader(text.splitlines(), dialect=dialect))

    def _load_xlsx(self, path: Path) -> list[dict[str, Any]]:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = (
                workbook[self.config.sheet]
                if self.config.sheet and self.config.sheet in workbook.sheetnames
                else workbook.worksheets[0]
            )
            rows = sheet.iter_rows(values_only=True)
            header = [str(cell) if cell is not None else "" for cell in next(rows, ())]
            return [
                dict(zip(header, values, strict=False))
                for values in rows
                if any(value is not None for value in values)
            ]
        finally:
            workbook.close()

    # -- Felder --------------------------------------------------------------

    def _field(self, row: dict[str, Any], name: str) -> Any:
        """Wert einer logischen Spalte ueber die konfigurierte Zuordnung."""
        column = self.config.columns.get(name)
        if not column:
            return None
        if column in row:
            return row[column]
        # Tolerant gegen Gross-/Kleinschreibung und Leerzeichen im Kopf.
        wanted = _normalize_key(column)
        for key, value in row.items():
            if _normalize_key(str(key)) == wanted:
                return value
        return None

    def _to_quote(self, row: dict[str, Any]) -> PriceQuote | None:
        name = _clean(self._field(row, "product_name"))
        if not name:
            return None
        amount = parse_amount(self._field(row, "amount"))
        currency = _clean(self._field(row, "currency")) or self.config.default_currency
        basis = _parse_basis(_clean(self._field(row, "basis")))
        if basis is PriceBasis.UNKNOWN and self.config.default_basis:
            # Ausdrueckliche Ansage in der Konfiguration ("diese Liste ist
            # netto") ist eine Angabe des Nutzers, keine Annahme des Tools.
            basis = PriceBasis(self.config.default_basis)
        vat_rate = _parse_rate(self._field(row, "vat_rate"))
        if vat_rate is None:
            vat_rate = self.config.default_vat_rate

        warnings: list[str] = []
        if amount is None:
            warnings.append("Kein Preis in der Liste")
        if basis is PriceBasis.UNKNOWN:
            warnings.append("Netto/Brutto nicht ausgewiesen")

        retrieved = parse_datetime(self._field(row, "price_date")) or (
            self._loaded_at or datetime.now(UTC)
        )
        return PriceQuote(
            supplier=_clean(self._field(row, "supplier")) or self.config.supplier or self.name,
            product_name=name,
            manufacturer=_clean(self._field(row, "manufacturer")),
            model_number=_clean(self._field(row, "model_number")),
            article_number=_clean(self._field(row, "article_number")),
            amount=amount,
            currency=currency,
            basis=basis,
            vat_rate=vat_rate,
            unit=_clean(self._field(row, "unit")),
            tiers=_parse_tiers(self._field(row, "tiers")),
            shipping_cost=parse_amount(self._field(row, "shipping_cost")),
            shipping_included=self.config.shipping_included,
            min_order_quantity=parse_amount(self._field(row, "min_order_quantity")),
            packaging_unit=parse_amount(self._field(row, "packaging_unit")),
            availability=_parse_availability(_clean(self._field(row, "availability"))),
            lead_time_days=int(days)
            if (days := parse_amount(self._field(row, "lead_time_days"))) is not None
            else None,
            url=_clean(self._field(row, "url")),
            retrieved_at=retrieved,
            warnings=warnings,
            provenance=Provenance(
                source=self.name,
                source_id=_clean(self._field(row, "article_number")),
                source_url=_clean(self._field(row, "url")),
                method="catalog",
                document=str(self.path.name),
                original_text="; ".join(
                    f"{key}={value}" for key, value in row.items() if _clean(value)
                )[:1000],
            ),
        )

    # -- Suche ---------------------------------------------------------------

    def _score_row(self, query: ProductQuery, quote: PriceQuote) -> float:
        """Vorauswahl in der Liste - die eigentliche Bewertung macht ``matching``.

        Hier geht es nur darum, aus zehntausend Zeilen die richtigen Kandidaten
        zu holen; welche davon wirklich passen, entscheidet danach das
        begruendete Matching.
        """
        if query.article_number and normalize_identifier(
            query.article_number
        ) == normalize_identifier(quote.article_number):
            return 1.0
        if query.model_number and normalize_identifier(query.model_number) == normalize_identifier(
            quote.model_number
        ):
            return 0.9
        return title_overlap(query.text, quote.product_name)

    async def search(self, query: ProductQuery) -> list[PriceQuote]:
        rows = self._load_rows()
        scored: list[tuple[float, PriceQuote]] = []
        for row in rows:
            quote = self._to_quote(row)
            if quote is None:
                continue
            score = self._score_row(query, quote)
            if score >= self.config.minimum_similarity:
                scored.append((score, quote))
        scored.sort(key=lambda pair: pair[0], reverse=True)
        return [quote for _score, quote in scored[: query.max_results]]

    async def health_check(self) -> PriceSourceStatus:
        """Liste einmal lesen und melden, wie viele Zeilen brauchbar sind."""
        try:
            rows = self._load_rows()
        except Exception as exc:  # noqa: BLE001 - der Health-Check meldet jeden Fehler
            return PriceSourceStatus(
                name=self.name,
                type=self.type_name,
                ok=False,
                message=f"{self.path}: {type(exc).__name__}: {exc}",
            )
        if not rows:
            return PriceSourceStatus(
                name=self.name,
                type=self.type_name,
                ok=False,
                message=f"{self.path}: keine Zeilen gelesen",
            )

        quotes = [quote for row in rows if (quote := self._to_quote(row)) is not None]
        with_price = sum(1 for quote in quotes if quote.has_price)
        calculable = sum(1 for quote in quotes if quote.net_amount()[0] is not None)
        message = (
            f"{len(rows)} Zeile(n), {len(quotes)} mit Bezeichnung, {with_price} mit Preis, "
            f"{calculable} kalkulationsfaehig (Netto ableitbar)"
        )
        if calculable == 0 and with_price:
            message += " | Hinweis: Netto/Brutto fehlt - 'default_basis' setzen"
        return PriceSourceStatus(
            name=self.name,
            type=self.type_name,
            ok=with_price > 0,
            message=message,
            sample_count=with_price,
        )
